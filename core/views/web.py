# -*- coding: utf-8 -*-

import string
import os
import time

from calendar import monthrange
from datetime import date, datetime, timedelta
from json import dumps

from django.http import HttpResponse
from django.conf import settings
from django.shortcuts import redirect, get_object_or_404
from django.core.paginator import Paginator, InvalidPage, EmptyPage
from django.utils.translation import ugettext_lazy as _
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.db.models import Q
from django.core.exceptions import PermissionDenied

from core import utils
from core.utils import render
from core.utils.special_days_ger import is_special_day, get_special_day_name
from core.jinja_lib import render_jinja
from core.models import (
    News, Company, Student, StudentGroup, Contact, Note,
    UserProfile, PresenceDay, PRESENCE_CHOICES
)
from core.forms import (
    NewsForm, SearchForm, StudentSearchForm, NoteForm,
    ProfileForm, NewUserForm, ExtendedSearchForm, StudentEditForm,
    ILBForm
)
from core.views import helper as h
from core.menu import menus
from barcode.codex import Code39
try:
    from barcode.writer import ImageWriter
except ImportError:
    ImageWriter = None
try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


# Create your views here.


def index(req):
    if req.user.is_authenticated():
        news_list = News.objects.all()
    else:
        news_list = News.objects.filter(public=True)
    paginator = Paginator(news_list, 5)
    try:
        page = int(req.GET.get('page', '1'))
    except ValueError:
        page = 1
    try:
        news = paginator.page(page)
    except (EmptyPage, InvalidPage):
        news = paginator.page(paginator.num_pages)
    ctx = dict(page_title=_(u'News'), menus=menus, news=news)
    return render_jinja(req, 'jinja_index.html', ctx)


@login_required
def internal_admin(req):
    if not req.user.is_superuser:
        raise PermissionDenied
    ctx = dict(page_title=_(u'Internal Admin Page'), menus=menus,
               need_ajax=True)
    return render(req, 'iadmin.html', ctx)


@login_required
def edit_profile(req):
    profile = req.user.userprofile
    if req.method == 'POST':
        form = ProfileForm(req.POST, instance=profile)
        form.save()
        messages.success(req, u'Alle Änderungen gespeichert.')
    else:
        form = ProfileForm(instance=profile)
    ctx = dict(page_title=_(u'My Profile'), menus=menus, form=form, dp=True)
    return render(req, 'colleagues/profile.html', ctx)


@login_required
def internal_phonelist(req):
    profiles = UserProfile.objects.select_related().filter(
        external=False
    ).exclude(user__username='admin').order_by('user__last_name')
    ctx = dict(page_title=_(u'Internal Phonelist'), menus=menus,
               profiles=profiles, dt=True)
    return render(req, 'colleagues/phonelist.html', ctx)


@login_required
def list_colleagues(req):
    internal = UserProfile.objects.select_related().filter(
        external=False
    ).exclude(user__username='admin').order_by('user__last_name')
    external = UserProfile.objects.select_related().filter(
        external=True
    ).exclude(user__username='admin').order_by('user__last_name')
    ctx = dict(page_title=_(u'Colleagues'), menus=menus, internal=internal,
               external=external)
    return render(req, 'colleagues/list.html', ctx)


@permission_required('auth.add_user', raise_exception=True)
def add_colleague(req):
    if req.method == 'POST':
        form = NewUserForm(req.POST)
        if form.is_valid():
            cd = form.cleaned_data
            uname = h.make_username(cd['lastname'], cd['firstname'])
            user = User.objects.create(
                username=uname, email=cd['email'],
                last_name=cd['lastname'], first_name=cd['firstname']
            )
            user.set_unusable_password()
            user.save()
            p = user.userprofile
            p.name_prefix = cd['name_prefix']
            p.phone = cd['phone']
            p.street = cd['street']
            p.zip_code = cd['zip_code']
            p.city = cd['city']
            p.country = cd['country']
            p.mobile = cd['mobile']
            p.birthdate = cd['birthdate']
            p.subjects = cd['subjects']
            p.can_login = cd['can_login']
            p.external = True
            p.save()
            messages.success(
                req, u'Neuer Datensatz (%(last)s, %(first)s) '
                     u'wurde gespeichert.' % {'last': user.last_name,
                                              'first': user.first_name}
            )
            return redirect('core-colleagues')
        else:
            messages.error(req, u'Bitte korrigieren Sie die Eingaben.')
    else:
        form = NewUserForm()
    ctx = dict(page_title=_(u'Add external'), menus=menus, form=form,
               dp=True)
    return render(req, 'colleagues/add.html', ctx)


@login_required
def filter_colleagues(req, filter):
    q = UserProfile.objects.select_related().exclude(user__username='admin')
    if filter == 'internal':
        q = q.filter(external=False)
        ptitle = _(u'Colleagues (internal): {0}')
    elif filter == 'external':
        q = q.filter(external=True)
        ptitle = _(u'Colleagues (external): {0}')
    else:
        ptitle = _(u'Colleagues (internal and external): {0}')
    ctx = dict(page_title=ptitle.format(q.count()), menus=menus, dt=True,
               coll=q.order_by('user__last_name'), filter=filter)
    return render(req, 'colleagues/filter.html', ctx)


@permission_required('auth.change_user', raise_exception=True)
def get_user_info(req, uid):
    profile = UserProfile.objects.select_related().get(user__id=int(uid))
    user = profile.user
    ctx = dict(u=user, p=profile)
    return render(req, 'colleagues/detail.html', ctx)


@permission_required('core.add_news')
def add_news(req):
    if req.method == 'POST':
        form = NewsForm(req.POST)
        if form.is_valid():
            news = form.save(commit=False)
            news.author = req.user
            news.save()
            messages.success(req, u'Meldung wurde hinzugefügt.')
            return redirect('/')
        messages.error(req, u'Bitte korrigieren Sie die falschen Felder.')
    else:
        form = NewsForm()
    ctx = dict(page_title=_(u'Add News'), menus=menus, form=form)
    return render(req, 'news/add.html', ctx)


@permission_required('core.add_note')
def add_note(req, id):
    contact = get_object_or_404(Contact, pk=int(id))
    company = contact.company
    old_notes = contact.notes.all().order_by('-date')
    if req.method == 'POST':
        form = NoteForm(req.POST)
        if form.is_valid():
            note = Note(subject=form.cleaned_data['subject'],
                        text=form.cleaned_data['text'])
            note.user = req.user
            note.contact = contact
            note.save()
            messages.success(req, u'Neue Bemerkung gespeichert.')
            return redirect('/core/companies/addnote/{0}/'.format(id))
        else:
            messages.error(req, u'Bitte Betreff und Text eingeben.')
    else:
        form = NoteForm()
    title = u'Neue Bemerkung: {name} ({company})'.format(
        name=contact.shortname(), company=company.short_name or company.name)
    ctx = dict(page_title=title, menus=menus, form=form,
               contact=contact, company=company, notes=old_notes)
    return render(req, 'companies/add_note.html', ctx)


def do_login(req):
    if req.method == 'POST':
        next_page = req.POST.get('next', '/')
        form = AuthenticationForm(req, req.POST)
        if form.is_valid():
            user = authenticate(
                username=form.cleaned_data['username'].lower(),
                password=form.cleaned_data['password']
            )
            if user is not None:
                p = user.userprofile
                if user.is_active and p.can_login:
                    login(req, user)
                    if user.is_superuser:
                        count = utils.remove_old_sessions()
                        messages.success(
                            req, u'{0} alte Sitzungen gelöscht'.format(count)
                        )
                    messages.success(req, u'Login akzeptiert.')
                    return redirect(next_page)
                else:
                    messages.error(req, u'Account ist deaktiviert.')
            else:
                messages.error(req, u'Benutzer und/oder Passwort falsch.')
    else:
        next_page = req.GET.get('next', '/')
        form = AuthenticationForm()
        req.session.set_test_cookie()
    ctx = dict(page_title=_(u'Login Page'), form=form, next_page=next_page)
    return render(req, 'login.html', ctx)


def do_logout(req):
    logout(req)
    messages.success(req, u'Erfolgreich abgemeldet.')
    return redirect('core-index')


@login_required
def company_details(req, company_id):
    company = Company.objects.get(pk=int(company_id))
    ctx = dict(page_title=company.name, menus=menus,
               c=company, single_view=True)
    return render(req, 'companies/view_company.html', ctx)


@login_required
def list_companies(req, startchar=''):
    if req.method == 'POST':
        form = SearchForm(req.POST)
        if form.is_valid():
            s = form.cleaned_data['search']
            companies = Company.objects.select_related().filter(
                Q(name__icontains=s) | Q(short_name__icontains=s))
        else:
            companies = Company.objects.none()
            messages.error(req, u'Ungültige Suche.')
    else:
        form = SearchForm()
        if not startchar:
            companies = Company.objects.select_related().all()
            for c in string.ascii_uppercase:
                companies = companies.exclude(name__istartswith=c)
        else:
            companies = Company.objects.select_related().filter(
                name__istartswith=startchar)
    companies = companies.exclude(**settings.EXCLUDE_FROM_COMPANY_LIST)
    ptitle = u'Firmen: {0} ({1})'.format(
        startchar or '-', companies.count()
    )
    ctx = dict(
        page_title=ptitle, companies=companies, menus=menus,
        chars=string.ascii_uppercase, form=form, single_view=False
    )
    return render(req, 'companies/list.html', ctx)


@login_required
def list_all_companies(req, only_with_students=False):
    q = Company.objects.select_related().all(
        ).exclude(**settings.EXCLUDE_FROM_COMPANY_LIST).order_by('name')
    if only_with_students:
        company_list = [x for x in q if x.has_students()]
        title = u'Firmen mit Azubis ({n})'
    else:
        company_list = list(q)
        title = u'Firmen ({n})'
    paginator = Paginator(company_list, 15)
    try:
        page = int(req.GET.get('page', '1'))
    except ValueError:
        page = 1
    try:
        companies = paginator.page(page)
    except (EmptyPage, InvalidPage):
        companies = paginator.page(paginator.num_pages)
    ctx = dict(
        page_title=title.format(n=len(company_list)),
        companies=companies, menus=menus,
        only_with_students=only_with_students, single_view=False,
        page=page, start=(page - 1) * 15 + 1
    )
    return render(req, 'companies/list_all.html', ctx)


@login_required
def list_students(req, startchar='', archive=False):
    if req.method == 'POST':
        form = StudentSearchForm(req.POST)
        form.fields['group'].choices = h.get_studentgroups()
        if form.is_valid():
            s = form.cleaned_data['search']
            q = (Q(lastname__istartswith=s) | Q(company__name__icontains=s)
                 | Q(barcode__istartswith=s) | Q(cabinet__icontains=s))
            if form.cleaned_data['group']:
                q &= Q(group__id=form.cleaned_data['group'])
            else:
                q |= Q(group__job_short__icontains=s)
            students = Student.objects.select_related().filter(q)
        else:
            students = Student.objects.none()
            messages.error(req, u'Ungültige Suche.')
    else:
        form = StudentSearchForm()
        form.fields['group'].choices = h.get_studentgroups()
        if not startchar:
            students = Student.objects.select_related().all()
            for c in string.ascii_uppercase:
                students = students.exclude(lastname__istartswith=c)
        else:
            students = Student.objects.select_related().filter(
                lastname__istartswith=startchar)
    students = [h.get_presence_details(x) for x in
                students.filter(finished=archive)]
    if archive:
        title = u'Azubis Archiv: {s} ({n})'
    else:
        title = u'Azubis: {s} ({n})'
    ctx = dict(
        page_title=title.format(s=startchar or '-', n=len(students)),
        students=students, menus=menus,
        archive=archive, startchar=startchar, chars=string.ascii_uppercase,
        form=form, need_ajax=True
    )
    return render(req, 'students/list.html', ctx)


@login_required
def list_all_students(req):
    student_list = Student.objects.select_related().all(
        ).order_by('lastname', 'group__job')
    paginator = Paginator(student_list, 30)
    try:
        page = int(req.GET.get('page', '1'))
    except ValueError:
        page = 1
    try:
        students = paginator.page(page)
    except (EmptyPage, InvalidPage):
        students = paginator.page(paginator.num_pages)
    ctx = dict(page_title=u'Alle Azubis ({0})'.format(
            students.paginator.count),
        menus=menus, students=students, page=page, start=(page - 1) * 30 + 1)
    return render(req, 'students/list_all.html', ctx)


@login_required
def search_student(req):
    students = None
    result = False
    if req.method == 'POST':
        result = True
        form = ExtendedSearchForm(req.POST)
        if form.is_valid():
            cd = form.cleaned_data
            search = Q(**{cd['search_for_1']: cd['search_1']})
            if cd['search_for_2'] and cd['search_2']:
                if cd['connect_with'] == u'and':
                    search &= Q(**{cd['search_for_2']: cd['search_2']})
                else:
                    search |= Q(**{cd['search_for_2']: cd['search_2']})
            students = Student.objects.select_related().filter(search)
    else:
        form = ExtendedSearchForm({'search_for_1': u'lastname'})
    ctx = dict(page_title=_(u'Search Student'), menus=menus,
               form=form, students=students, result=result, need_ajax=True)
    return render(req, 'students/search.html', ctx)


@permission_required('core.change_student')
def edit_student(req, sid):
    s = Student.objects.get(id=int(sid))
    initial = dict(cabinet=s.cabinet, key=s.key, exam_1=s.exam_1,
                   exam_2=s.exam_2, finished=s.finished)
    form = StudentEditForm(initial=initial)
    ctx = dict(student=s, form=form)
    return render(req, 'students/edit_short.html', ctx)


@login_required
def list_groups(req):
    g = StudentGroup.objects.select_related().all()
    groups = [x for x in g if not x.finished()]
    ctx = dict(page_title=_(u'Groups - Overview'), groups=groups, menus=menus,
               dt=True)
    return render(req, 'students/groups.html', ctx)


@login_required
def group_details(req, gid):
    try:
        gid = int(gid)
    except ValueError:
        messages.error(req, u'Gruppen-ID muss eine Zahl sein.')
        return redirect('core-groups')
    try:
        group = StudentGroup.objects.select_related().get(pk=gid)
    except StudentGroup.DoesNotExist:
        messages.error(req, u'Gruppe (ID: %d) existiert nicht.' % gid)
        return redirect('core-groups')
    if group.school_nr:
        ptitle = _(u'Group - {name} ({nr})'.format(
            name=group.name(), nr=group.school_nr))
    else:
        ptitle = _(u'Group - {name}'.format(name=group.name()))
    ctx = dict(page_title=ptitle, group=group, menus=menus, need_ajax=True)
    return render(req, 'students/group_detail.html', ctx)


@login_required
def presence_overview(req):
    d = date.today()
    month = d.month
    year = d.year
    last_month = month - 1 if month > 1 else 12
    lyear = year if last_month != 12 else year - 1
    q = Q(date__month=month) & Q(date__year=year)
    q |= Q(date__month=last_month) & Q(date__year=lyear)
    q2 = Q(entry__isnull=True) | Q(entry__exact=u'')
    jobs = StudentGroup.objects.values_list('job', flat=True)
    jobs = list(set(jobs))
    jobs.sort()
    groups = []
    for j in jobs:
        gr = StudentGroup.objects.select_related().filter(job=j).order_by(
            'start_date')
        gr = [x for x in gr if not x.finished()]
        for g in gr:
            g.pdfs = g.presence_printouts.filter(q).order_by('-date')
            g.pdays = PresenceDay.objects.filter(
                student__group=g, date__month=last_month, date__year=lyear
            ).exclude(q2).count()
        groups.append((j, gr))
    ctx = dict(page_title=_(u'Presence Overview'), groups=groups, menus=menus,
               month=last_month, jobs=jobs, dp=True, dt=True)
    return render(req, 'presence/overview.html', ctx)


@login_required
def presence_for_group(req, gid):
    if req.method == 'POST':
        start = req.POST['start'] or None
        end = req.POST['end'] or None
    else:
        start = end = None
    try:
        gid = int(gid)
    except ValueError:
        messages.error(req, u'Gruppen-ID muss eine Zahl sein.')
        return redirect('core-presence')
    try:
        group = StudentGroup.objects.select_related().get(pk=gid)
    except StudentGroup.DoesNotExist:
        messages.error(req, u'Gruppe (ID: %d) existiert nicht.' % gid)
        return redirect('core-presence')
    _d = date.today()
    d = _d - timedelta(days=7)
    start = utils.get_date(start, d)
    end = utils.get_date(end, _d)
    if start > end:
        messages.error(req, u'Das Enddatum liegt vor dem Startdatum!')
        return redirect('core-presence')
    req.session['presence_start'] = start.strftime('%Y-%m-%d')
    req.session['presence_end'] = end.strftime('%Y-%m-%d')
    dt = end - start
    _students = h.sort_students_for_presence(group.students)
    students = h.get_presence(_students, start, end)
    days = (start + timedelta(days=x) for x in xrange(dt.days + 1))
    ctx = dict(
        page_title=u'Anwesenheit für Gruppe {0}'.format(unicode(group)),
        group=group, students=students, menus=menus, start=start, end=end,
        days=[x for x in days if x.weekday() not in (5, 6)],
        choices=[x[0] for x in PRESENCE_CHOICES], legend=PRESENCE_CHOICES[1:],
        today=date.today(), dt=True, need_ajax=True
    )
    return render(req, 'presence/group.html', ctx)


@permission_required('core.change_presenceday')
def presence_edit(req, student_id):
    _student = Student.objects.select_related().get(id=int(student_id))
    today = date.today().strftime('%Y-%m-%d')
    _start = req.session.get('presence_start', today)
    _end = req.session.get('presence_end', today)
    start = datetime.strptime(_start, '%Y-%m-%d').date()
    end = datetime.strptime(_end, '%Y-%m-%d').date()
    student, days = h.get_presence([_student], start, end)[0]
    from_ = req.session.get('from', 'daily')
    req.session['from'] = 'daily'
    ctx = dict(
        page_title=u'Anwesenheit - {0}'.format(unicode(student)),
        menus=menus, student=student, days=days, start=start, end=end,
        choices=PRESENCE_CHOICES, dt=True, need_ajax=True, from_=from_
    )
    return render(req, 'presence/edit.html', ctx)


@login_required
def presence_printouts(req, job):
    groups = StudentGroup.objects.select_related().filter(
        job=job).order_by('-start_date', 'job_short')
    jobs = StudentGroup.objects.values_list('job', flat=True)
    jobs = list(set(jobs))
    jobs.sort()
    ctx = dict(page_title=u'Anwesenheiten {0}'.format(job),
               menus=menus, jobs=jobs, groups=groups)
    return render(req, 'presence/list_printouts.html', ctx)


@login_required
def presence_for_student_overview(req, student_id):
    student = Student.objects.select_related().get(id=int(student_id))
    student = h.get_presence_details(student)
    q = Q(entry__in=[u'T', u'F', u'K', u'|', u'U']) | Q(lateness__gt=0)
    days = student.presence_days.select_related().filter(q).order_by('date')
    ctx = dict(page_title=u'Anwesenheitsübersicht', subtitle=unicode(student),
               menus=menus, student=student, days=days, dt=True)
    return render(req, 'students/presence_overview.html', ctx)


@login_required
def get_next_birthdays(req):
    days = int(req.GET.get('days', '14'))
    choice = [7, 14, 30, 90, 180]
    start = date.today()
    today = (start.month, start.day)
    start = start - timedelta(days=10)
    dates = [start + timedelta(days=x) for x in xrange(days)]
    users = []
    students = []
    for d in dates:
        for p in UserProfile.objects.select_related(
          ).filter(birthdate__month=d.month, birthdate__day=d.day):
            p.bsclass = utils.get_birthday_color(
                (p.birthdate.month, p.birthdate.day), today)
            p.bdate = d
            users.append(p)
        for s in Student.objects.select_related(
          ).filter(birthdate__month=d.month, birthdate__day=d.day):
            s.bsclass = utils.get_birthday_color(
                (s.birthdate.month, s.birthdate.day), today)
            s.bdate = d
            students.append(s)
    ctx = dict(
        page_title=_(u'Birthdays, next {0} days'.format(days)),
        menus=menus, users=users, students=students, choice=choice, days=days,
        today=start
    )
    return render(req, 'colleagues/birthdays.html', ctx)


@login_required
def export_group_excel(req, gid):
    if Workbook is None:
        return utils.error(u'Die Excelerweiterung ist nicht installiert.')
    group = StudentGroup.objects.select_related().get(id=int(gid))
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.cell('A1').value = unicode(group)
    ws.cell('A3').value = u'Firma'
    ws.cell('B3').value = u'Name'
    ws.cell('C3').value = u'Vorname'
    row = 4
    for s in group.students.filter(finished=False).order_by(
      'company__short_name', 'lastname'):
        if s.company:
            comp = s.company.short_name
        else:
            comp = u'-'
        ws.cell('A{0}'.format(row)).value = comp
        ws.cell('B{0}'.format(row)).value = s.lastname
        ws.cell('C{0}'.format(row)).value = s.firstname
        ws.cell('D{0}'.format(row)).value = s.email
        row += 1
    dest = os.path.join(
        settings.LATEX_BUILD_DIR, 'excel_exp_{0}.xlsx'.format(time.time())
    )
    wb.save(dest)
    with open(dest, 'rb') as fp:
        response = HttpResponse(
            fp.read(), content_type='application/vnd.'
            'openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = 'attachment; filename="{0}.xlsx"'.format(
        group.name())
    os.remove(dest)
    return response


@login_required
def export_group_json(req, gid):
    group = StudentGroup.objects.select_related().get(id=int(gid))
    exp = dict(job=group.job, name=group.name())
    students = []
    for s in group.students.filter(finished=False).order_by('lastname'):
        students.append({
            'lastname': s.lastname,
            'firstname': s.firstname,
            'birthdate': s.birthdate.strftime('%Y-%m-%d'),
            'barcode': s.barcode,
        })
    exp['students'] = students
    response = HttpResponse(dumps(exp), content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="{0}.json"'.format(
        group.name()
    )
    return response


@login_required
def export_group_moodle(req, gid):
    group = StudentGroup.objects.select_related().get(id=int(gid))
    fields = [u'username', u'lastname', u'firstname', u'email', u'cohort1']
    response = HttpResponse(content_type='text/csv')
    writer = utils.UnicodeCSVWriter(response, fields)
    writer.writeheader()
    cohort = group.job_short.replace(u' ', u'').lower()
    if group.suffix:
        cohort = u'{0}{1}'.format(cohort, group.suffix)
    for s in group.students.filter(finished=False).order_by('lastname'):
        if s.email.strip():
            writer.writerow({
                'username': s.email.lower(),
                'lastname': s.lastname,
                'firstname': s.firstname,
                'email': s.email.lower(),
                'cohort1': cohort,
            })
    response['Content-Disposition'] = 'attachment; filename={0}.csv'.format(
        cohort
    )
    return response


def barcode(req, format, barcode=''):
    if format == 'svg':
        bc = Code39(barcode, add_checksum=False)
        mimetype = 'image/svg+xml'
    else:
        if ImageWriter is None:
            return utils.error(req, _('PIL is not installed.'))
        bc = Code39(barcode, ImageWriter(), add_checksum=False)
        mimetype = 'image/{0}'.format(format)
    response = HttpResponse(mimetype=mimetype)
    try:
        bc.write(response, options={'format': format.upper()})
    except KeyError:
        return utils.error(req, _('Unsupported barcode format.'))
    return response


# Functions for user defined presence list (one per user)

@login_required
def select_groups(req):
    jobs = StudentGroup.objects.values_list('job', flat=True)
    jobs = list(set(jobs))
    jobs.sort()
    groups = []
    for j in jobs:
        g = StudentGroup.objects.filter(job=j).order_by('-start_date')
        groups.append((j, [x for x in g if x.active_count()]))
    profile = req.user.userprofile
    config = profile.config()
    pgroups = config.get('pgroups', [])
    students = Student.objects.select_related().filter(
        group__id__in=pgroups, finished=False).order_by(
            'group__job_short', 'lastname')
    ctx = dict(page_title=_(u'Edit own presence list'), menus=menus, jobs=jobs,
               groups=groups, need_ajax=True, students=students)
    return render(req, 'presence/select_groups.html', ctx)


@login_required
def mystudents(req):
    profile = req.user.userprofile
    config = profile.config()
    pgroups = config.get('pgroups', [])
    students = Student.objects.select_related().filter(
        group__id__in=pgroups, finished=False).order_by(
            'group__job_short', 'lastname')
    return render(req, 'presence/studentlist.html', {'students': students})


@login_required
def mypresence(req):
    if req.method == 'POST':
        start = req.POST['start'] or None
        end = req.POST['end'] or None
    else:
        start = end = None
    _d = date.today()
    d = _d - timedelta(days=7)
    start = utils.get_date(start, d)
    end = utils.get_date(end, _d)
    if start > end:
        messages.error(req, u'Das Enddatum liegt vor dem Startdatum!')
        return redirect('core-presence')
    req.session['presence_start'] = start.strftime('%Y-%m-%d')
    req.session['presence_end'] = end.strftime('%Y-%m-%d')
    dt = end - start
    _studs = h.get_students(req.user)
    _students = h.sort_students_for_presence(_studs)
    students = h.get_presence(_students, start, end)
    days = (start + timedelta(days=x) for x in xrange(dt.days + 1))
    ctx = dict(
        page_title=_(u'My Presence'),
        students=students, menus=menus, start=start, end=end,
        days=[x for x in days if x.weekday() not in (5, 6)],
        choices=[x[0] for x in PRESENCE_CHOICES], legend=PRESENCE_CHOICES[1:],
        today=date.today(), dt=True, need_ajax=True
    )
    return render(req, 'presence/group.html', ctx)


@permission_required('core.change_studentgroup')
def make_ilb_group(req, gid):
    group = StudentGroup.objects.select_related().get(id=int(gid))
    step = 1
    days = None
    cd = None
    day_count = 0
    if req.method == 'POST':
        form = ILBForm(req.POST)
        if form.is_valid():
            step = 2
            cd = form.cleaned_data
            req.session['selected_students'] = cd['students']
            dt = cd['end'] - cd['start']
            cd['course'] = h.check_course(cd['course'], group)
            not_select = [cd['school1'], cd['school2'], 5, 6]
            days = []
            for day in xrange(dt.days + 1):
                d = cd['start'] + timedelta(days=day)
                if d.weekday() == 6:
                    continue
                day_name = ''
                if is_special_day(d):
                    day_name = get_special_day_name(d)
                if d.weekday() in not_select or day_name:
                    days.append((False, d, day_name))
                else:
                    day_count += 1
                    days.append((True, d, day_name))
    else:
        students = group.students.filter(
            finished=False).order_by('company', 'lastname')
        form = ILBForm()
        form.fields['students'].choices = [
            (x.id, x.fullname()) for x in students]
    ctx = dict(page_title=u'ILB Liste erstellen', group=group, menus=menus,
               form=form, dp=True, step=step, days=days, need_ajax=True,
               cd=cd, day_count=day_count)
    return render(req, 'students/ilb_setup.html', ctx)


@login_required
def show_monthly_presence(req, year=None, month=None):
    today = date.today()
    pre = today - timedelta(days=today.day + 1)
    if year is None:
        year = pre.year
    if month is None:
        month = pre.month
    year, month = int(year), int(month)
    start = date(year, month, 1)
    days_in_month = monthrange(year, month)[1]
    end = date(year, month, days_in_month)
    req.session['presence_start'] = start.strftime('%Y-%m-%d')
    req.session['presence_end'] = end.strftime('%Y-%m-%d')
    req.session['from'] = 'monthly'
    students, days = h.get_students_for_month(start, end)
    next = start + timedelta(days=start.day + days_in_month)
    prev = start - timedelta(days=start.day + 1)
    ctx = dict(page_title=u'Anwesenheiten im {0} {1}'.format(
               utils.GERMAN_MONTH[month], year),
               students=students.order_by('lastname', 'firstname'),
               days=days, dt=True, menus=menus, prev=prev, next=next)
    return render(req, 'presence/monthly.html', ctx)


@permission_required('core.add_note')
def notes_overview(req, show=u'short'):
    query = Q(contacts__isnull=False)
    sub = u'Alle'
    if show == u'short':
        query &= Q(contacts__notes__isnull=False)
        sub = u'Notizen'
    companies = Company.objects.select_related().filter(
        query).distinct().order_by('name')
    ctx = dict(page_title=u'Firmenkontakte', subtitle=sub,
               companies=companies, menus=menus, show=show)
    return render(req, 'companies/notes.html', ctx)


@permission_required('core.delete_studentgroup')
def delete_group(req, gid):
    group = StudentGroup.objects.select_related().get(id=int(gid))
    gname = unicode(group)
    if req.method == 'POST':
        group.presence_printouts.all().delete()
        deleted = []
        for stud in group.students.all():
            try:
                _id = stud.id
                name, count = h.delete_one_student(_id, True)
                deleted.append(_id)
            except Exception as e:
                messages.error(
                    req, u'Beim Löschen eines Teilnehmers ist ein Fehler '
                    u'aufgetreten: {0}'.format(e)
                )
        try:
            group.delete()
            req.session['group_name'] = gname
            req.session['deleted'] = deleted
            return redirect('core-group-delete-success')
        except Exception as e:
            messages.error(
                req, u'Beim Löschen der Gruppe ist ein Fehler'
                u'aufgetreten: {0}'.format(e)
            )
    ctx = dict(page_title=u'Gruppe löschen', group=group, menus=menus,
               subtitle=gname)
    return render(req, 'students/delete_group.html', ctx)


def delete_group_success(req):
    gname = req.session.pop('group_name')
    ids = req.session.pop('deleted')
    ctx = dict(page_title=u'DSGVO', gname=gname,
               ids=ids, menus=menus)
    return render(req, 'students/delete_group_success.html', ctx)
